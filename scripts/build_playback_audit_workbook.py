#!/usr/bin/env python3
"""Build a polished Google Sheets-compatible XLSX from the live audit JSON."""

from __future__ import annotations

import argparse
import csv
import json
import statistics
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


GREEN = "C6EFCE"
GREEN_TEXT = "006100"
RED = "FFC7CE"
RED_TEXT = "9C0006"
BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
GOLD = "FFF2CC"
WHITE = "FFFFFF"
GRAY = "E7E6E6"
THIN = Side(style="thin", color="D9E1F2")


def type_label(kind: str) -> str:
    return {
        "event": "Today Match",
        "channel": "TV Channel",
        "movie": "Movie",
    }.get(kind, kind)


def local_time(value: str) -> str:
    if not value:
        return ""
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    return parsed.astimezone(timezone(timedelta(hours=6))).strftime("%Y-%m-%d %H:%M:%S")


def percentile(values: list[float], ratio: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    return ordered[round((len(ordered) - 1) * ratio)]


def attempt_count(row: dict) -> int:
    return 1 + len(row.get("attempt_history") or [])


def first_attempt(row: dict) -> dict:
    history = row.get("attempt_history") or []
    return history[0] if history else row


def resolution(row: dict) -> str:
    diag = row.get("diagnostic") or {}
    width = int(diag.get("video_width") or 0)
    height = int(diag.get("video_height") or 0)
    return f"{width}x{height}" if width and height else ""


def network_failures(row: dict) -> str:
    failures = (row.get("diagnostic") or {}).get("network_failures") or []
    return " | ".join(dict.fromkeys(str(value) for value in failures))


def observed_state(row: dict) -> str:
    diag = row.get("diagnostic") or {}
    message = str(diag.get("player_message") or "").strip()
    if message:
        return message
    return row.get("browser_evidence") or ""


def detail_row(row: dict, serial: int) -> list:
    first = first_attempt(row)
    load = row.get("load_time_seconds")
    return [
        serial,
        row.get("category", ""),
        row.get("name", ""),
        row.get("status", ""),
        float(load) if load not in (None, "") else None,
        attempt_count(row),
        first.get("status", ""),
        row.get("problem", ""),
        observed_state(row),
        resolution(row),
        row.get("browser_evidence", ""),
        network_failures(row),
        local_time(row.get("tested_at", "")),
        row.get("uid", ""),
    ]


DETAIL_HEADERS = [
    "Serial", "Category", "Name", "Final Status", "Load Time (seconds)",
    "Attempts", "First Attempt", "Final Problem", "Player/UI Message",
    "Decoded Resolution", "Browser Evidence", "Sanitized Network Failures",
    "Final Test Time (Dhaka)", "Catalogue UID",
]


def style_title(ws, title: str, subtitle: str, last_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    cell = ws.cell(1, 1, title)
    cell.font = Font(size=18, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).font = Font(italic=True, color="404040")
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34


def style_header(row) -> None:
    for cell in row:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)


def add_table(ws, start_row: int, end_row: int, end_col: int, name: str) -> None:
    if end_row <= start_row:
        return
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def build_detail_sheet(wb: Workbook, name: str, rows: list[dict], subtitle: str) -> None:
    ws = wb.create_sheet(name)
    style_title(ws, name, subtitle, len(DETAIL_HEADERS))
    ws.append([])
    ws.append(DETAIL_HEADERS)
    style_header(ws[4])
    for serial, row in enumerate(rows, 1):
        ws.append(detail_row(row, serial))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:N{ws.max_row}"
    add_table(ws, 4, ws.max_row, len(DETAIL_HEADERS), f"{name.replace(' ', '')}Audit")
    widths = [9, 18, 42, 13, 19, 10, 15, 48, 36, 20, 54, 52, 23, 34]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.column_dimensions["N"].hidden = True
    for row in ws.iter_rows(min_row=5):
        row[0].alignment = Alignment(horizontal="center")
        row[3].alignment = Alignment(horizontal="center")
        row[4].number_format = "0.000"
        row[4].alignment = Alignment(horizontal="center")
        row[5].alignment = Alignment(horizontal="center")
        row[6].alignment = Alignment(horizontal="center")
        for cell in row[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.conditional_formatting.add(
        f"D5:D{ws.max_row}",
        CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_TEXT, bold=True)),
    )
    ws.conditional_formatting.add(
        f"D5:D{ws.max_row}",
        CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TEXT, bold=True)),
    )
    ws.sheet_view.showGridLines = False


def build_summary(wb: Workbook, payload: dict, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Summary"
    style_title(
        ws,
        "Click TV Full Live Playback Audit",
        "Real desktop Google Chrome UI clicks. PASS requires a decoded video frame plus measurable playback progress. Raw/tokenized source URLs are excluded.",
        8,
    )
    counts = Counter(row["status"] for row in rows)
    events = [row for row in rows if row["kind"] == "event"]
    channels = [row for row in rows if row["kind"] == "channel"]
    movies = [row for row in rows if row["kind"] == "movie"]
    loads = [float(row["load_time_seconds"]) for row in rows if row["status"] == "PASS"]
    times = sorted(row.get("tested_at", "") for row in rows if row.get("tested_at"))
    recoveries = [row for row in rows if row["status"] == "PASS" and first_attempt(row).get("status") != "PASS"]

    metrics = [
        ("Site", payload.get("base_url", "")),
        ("Test window (Dhaka)", f"{local_time(times[0])} to {local_time(times[-1])}" if times else ""),
        ("Total catalogue items", len(rows)),
        ("Today matches", len(events)),
        ("TV channels", len(channels)),
        ("Movies", len(movies)),
        ("Final PASS", counts.get("PASS", 0)),
        ("Final FAIL", counts.get("FAIL", 0)),
        ("Missing / skipped", len(payload.get("inventory", [])) - len(rows)),
        ("Automation ERROR", counts.get("ERROR", 0)),
        ("Recovered on retry", len(recoveries)),
        ("Average PASS load time", round(statistics.mean(loads), 3) if loads else 0),
        ("Median PASS load time", round(statistics.median(loads), 3) if loads else 0),
        ("95th percentile PASS load", round(percentile(loads, 0.95), 3) if loads else 0),
        ("Maximum PASS load time", round(max(loads), 3) if loads else 0),
    ]
    ws.append([])
    ws.append(["Metric", "Value"])
    style_header(ws[4][:2])
    for metric, value in metrics:
        ws.append([metric, value])
    for cell in ws[4]:
        cell.border = Border(bottom=THIN)
    ws[5][1].hyperlink = payload.get("base_url", "")
    ws[5][1].style = "Hyperlink"

    category_start = ws.max_row + 3
    ws.cell(category_start, 1, "Category Summary")
    ws.cell(category_start, 1).font = Font(size=14, bold=True, color=BLUE)
    headers = ["Type", "Category", "Total", "PASS", "FAIL", "Pass Rate", "Avg PASS Load (s)", "Median PASS Load (s)"]
    ws.append(headers)
    style_header(ws[ws.max_row])
    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in rows:
        grouped[(row["kind"], row["category"])].append(row)
    for (kind, category), items in grouped.items():
        passed = [row for row in items if row["status"] == "PASS"]
        category_loads = [float(row["load_time_seconds"]) for row in passed]
        ws.append([
            type_label(kind),
            category,
            len(items),
            len(passed),
            sum(row["status"] == "FAIL" for row in items),
            len(passed) / len(items) if items else 0,
            round(statistics.mean(category_loads), 3) if category_loads else None,
            round(statistics.median(category_loads), 3) if category_loads else None,
        ])
    category_header = category_start + 1
    add_table(ws, category_header, ws.max_row, 8, "CategorySummary")
    for cell in ws[category_header + 1:ws.max_row + 1]:
        cell[5].number_format = "0.0%"
        cell[6].number_format = "0.000"
        cell[7].number_format = "0.000"

    notes_start = ws.max_row + 3
    ws.cell(notes_start, 1, "Interpretation Notes")
    ws.cell(notes_start, 1).font = Font(size=14, bold=True, color=BLUE)
    notes = [
        "PASS means Chrome decoded video frames and currentTime advanced after clicking the catalogue card.",
        "FAIL rows were tested twice unless the first attempt passed; Attempts shows the actual count.",
        "Load Time is reported only for successful playback. A failed stream has no valid load time.",
        "ORB failures usually mean the browser received a blocked, non-media, or invalid cross-origin response from the media/proxy route.",
        "The audit records every live catalogue card exactly once in the final result; no Today Match, channel, or movie was skipped.",
    ]
    for note in notes:
        ws.append(["•", note])
    ws.column_dimensions["A"].width = 29
    ws.column_dimensions["B"].width = 46
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 19
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


def build_problem_summary(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Problem Summary")
    style_title(ws, "Problem Summary", "Final failure reasons after retry, grouped by problem and content type.", 7)
    ws.append([])
    headers = ["Problem", "Total", "Today Matches", "TV Channels", "Movies", "Share of All FAIL", "Example Items"]
    ws.append(headers)
    style_header(ws[4])
    failed = [row for row in rows if row["status"] == "FAIL"]
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in failed:
        grouped[row.get("problem") or "Unclassified failure"].append(row)
    for problem, items in sorted(grouped.items(), key=lambda pair: (-len(pair[1]), pair[0])):
        examples = "; ".join(row["name"] for row in items[:8])
        ws.append([
            problem,
            len(items),
            sum(row["kind"] == "event" for row in items),
            sum(row["kind"] == "channel" for row in items),
            sum(row["kind"] == "movie" for row in items),
            len(items) / len(failed) if failed else 0,
            examples,
        ])
    add_table(ws, 4, ws.max_row, 7, "ProblemSummary")
    for row in ws.iter_rows(min_row=5):
        row[5].number_format = "0.0%"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [62, 12, 18, 16, 12, 18, 85]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def build_recoveries(wb: Workbook, rows: list[dict]) -> None:
    recovered = [row for row in rows if row["status"] == "PASS" and first_attempt(row).get("status") != "PASS"]
    ws = wb.create_sheet("Retry Recoveries")
    style_title(ws, "Retry Recoveries", "Items that failed initially but played during the longer visible-Chrome retry.", 7)
    ws.append([])
    headers = ["Type", "Category", "Name", "Final Load Time (s)", "Attempts", "First Problem", "Final Evidence"]
    ws.append(headers)
    style_header(ws[4])
    for row in recovered:
        first = first_attempt(row)
        ws.append([
            type_label(row["kind"]),
            row["category"], row["name"], float(row["load_time_seconds"]), attempt_count(row),
            first.get("problem", ""), row.get("browser_evidence", ""),
        ])
    add_table(ws, 4, ws.max_row, 7, "RetryRecoveries")
    widths = [15, 18, 42, 20, 10, 55, 55]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=5):
        row[3].number_format = "0.000"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def export_google_csv(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Type"] + DETAIL_HEADERS)
        for serial, row in enumerate(rows, 1):
            writer.writerow([type_label(row["kind"])] + detail_row(row, serial))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input")
    parser.add_argument("output")
    parser.add_argument("--csv-output", default="")
    args = parser.parse_args()
    input_path = Path(args.input).resolve()
    output_path = Path(args.output).resolve()
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    rows = payload.get("results", [])
    inventory = payload.get("inventory", [])
    keys = [(row["category_id"], row["uid"]) for row in rows]
    if len(rows) != len(inventory) or len(keys) != len(set(keys)):
        raise RuntimeError(
            f"Incomplete or duplicate audit: inventory={len(inventory)} rows={len(rows)} unique={len(set(keys))}"
        )
    if any(row.get("status") not in {"PASS", "FAIL"} for row in rows):
        raise RuntimeError("Audit still contains ERROR or unknown status rows")

    wb = Workbook()
    build_summary(wb, payload, rows)
    build_problem_summary(wb, rows)
    build_detail_sheet(
        wb, "Today Matches", [row for row in rows if row["kind"] == "event"],
        "Every Today Match card tested from the deployed site in visible desktop Chrome.",
    )
    build_detail_sheet(
        wb, "TV Channels", [row for row in rows if row["kind"] == "channel"],
        "All live TV catalogue cards. Filter Final Status, Category, load time, or problem.",
    )
    build_detail_sheet(
        wb, "Movies", [row for row in rows if row["kind"] == "movie"],
        "All movie catalogue cards. Filter Final Status, Category, load time, or problem.",
    )
    build_recoveries(wb, rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Reopen to prove that the produced XLSX package is readable and complete.
    verified = load_workbook(output_path, read_only=True, data_only=False)
    expected = {"Summary", "Problem Summary", "Today Matches", "TV Channels", "Movies", "Retry Recoveries"}
    if set(verified.sheetnames) != expected:
        raise RuntimeError(f"Workbook sheet verification failed: {verified.sheetnames}")
    expected_counts = Counter(row["kind"] for row in rows)
    if (
        verified["Today Matches"].max_row - 4 != expected_counts.get("event", 0)
        or verified["TV Channels"].max_row - 4 != expected_counts.get("channel", 0)
        or verified["Movies"].max_row - 4 != expected_counts.get("movie", 0)
    ):
        raise RuntimeError("Workbook row-count verification failed")
    verified.close()

    if args.csv_output:
        export_google_csv(Path(args.csv_output).resolve(), rows)
    print(
        f"WORKBOOK_OK path={output_path} rows={len(rows)} "
        f"matches={expected_counts.get('event', 0)} channels={expected_counts.get('channel', 0)} "
        f"movies={expected_counts.get('movie', 0)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
